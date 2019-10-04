from pathlib import Path
import openpyxl as xl
path_1 = Path()
for file in path_1.glob('*.xlsx'):
    wb = xl.load_workbook(file)
    sheet = wb['acounts']
    for column in range(1, sheet.max_column + 1):
        for row in range(1, sheet.max_row + 1):
            cell_x = sheet.cell(row, column)
            cell_x.value = 'HERO'
    wb.save(file)
